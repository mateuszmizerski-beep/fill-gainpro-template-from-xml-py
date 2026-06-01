#!/usr/bin/env python3
"""Fill a Gain.pro-style Excel financial template from a Polish ESF XML file.

The script was calibrated on the Valvex 2024 XML / 2023-2024 workbook case.
It fills only the raw input rows that can be sourced reliably from the XML and
leaves formula rows for Excel to recalculate.
"""

from __future__ import annotations

import argparse
from copy import copy
import re
import sys
import xml.etree.ElementTree as ET
from dataclasses import dataclass
from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Iterable

try:
    from openpyxl import load_workbook
    from openpyxl.comments import Comment
    from openpyxl.utils import get_column_letter
except ImportError as exc:  # pragma: no cover - user environment guard
    raise SystemExit(
        "Missing dependency: openpyxl. Install it with `pip install openpyxl`."
    ) from exc


HEADER_ROW = 2
FINANCIALS_SHEET = "Financials"
PLN_TO_PLNM = Decimal("1000000")
CURRENT_PERIOD_TAG = "KwotaA"
CAGR_COLUMN = 21  # Column U
MIN_ANNUALISATION_DAYS = 90
FULL_YEAR_DAY_COUNTS = {365, 366}
TRANSFORMED_COMPARATIVE_TAG = "KwotaB1"
# In three-column ESF files, the far-right comparative column is KwotaB1.
# If KwotaB1 is missing or is an all-zero transformed column, fall back to KwotaB.
COMPARATIVE_PERIOD_TAGS = (TRANSFORMED_COMPARATIVE_TAG, "KwotaB")
AMOUNT_TAGS = {"KwotaA", "KwotaB", "KwotaB1"}
PRESERVED_COLUMNS = {1, 2, 3, 4, 20, 21}  # A:D, T:U
STATEMENT_ROOT_TAGS = {
    "JednostkaInna",
    "JednostkaMala",
    "JednostkaMikro",
    "JednostkaOp",
}
STATEMENT_SECTION_TAGS = {"Bilans", "RZiS", "RachPrzeplywow"}


@dataclass(frozen=True)
class Mapping:
    row_label: str
    xml_paths: tuple[str, ...]
    section_after: str | None = None
    section_before: str | None = None
    row_label_fallbacks: tuple[str, ...] = ()
    write_label: str | None = None
    confidence_label: str | None = None
    scale: Decimal = PLN_TO_PLNM
    blank_if_zero: bool = True


@dataclass(frozen=True)
class FillJob:
    xml_data: "XmlFinancials"
    year: int
    period_tags: tuple[str, ...]
    period_xml_data: "XmlFinancials | None" = None


MAPPINGS: tuple[Mapping, ...] = (
    # Reported figures
    Mapping("Other income", ("RZiS/RZiSKalk/G", "RZiS/RZiSPor/D"), "1. REPORTED FIGURES", "2. ADJUSTMENTS "),
    Mapping("COGS", ("RZiS/RZiSKalk/B",), "1. REPORTED FIGURES", "2. ADJUSTMENTS ", confidence_label="Gross margin"),
    Mapping(
        "Reported EBIT",
        ("RZiS/RZiSKalk/I", "RZiS/RZiSPor/F"),
        "1. REPORTED FIGURES",
        "2. ADJUSTMENTS ",
        confidence_label="EBIT",
    ),
    Mapping(
        "D&A",
        ("RachPrzeplywow/PrzeplywyPosr/A/A_II/A_II_1",),
        "1. REPORTED FIGURES",
        "2. ADJUSTMENTS ",
        confidence_label="EBITDA",
    ),
    Mapping(
        "Fixed assets",
        ("Bilans/Aktywa/Aktywa_A",),
        "1. REPORTED FIGURES",
        "2. ADJUSTMENTS ",
        confidence_label="Total assets",
    ),
    Mapping(
        "Current assets",
        ("Bilans/Aktywa/Aktywa_B",),
        "1. REPORTED FIGURES",
        "2. ADJUSTMENTS ",
        confidence_label="Total assets",
    ),
    Mapping(
        "Stocks / inventories",
        ("Bilans/Aktywa/Aktywa_B/Aktywa_B_I",),
        "1. REPORTED FIGURES",
        "2. ADJUSTMENTS ",
        confidence_label="Inventory",
    ),
    Mapping(
        "Trade debtors / receivables",
        (
            "Bilans/Aktywa/Aktywa_B/Aktywa_B_II/Aktywa_B_II_3/Aktywa_B_II_3_A",
            "Bilans/Aktywa/Aktywa_B/Aktywa_B_II/Aktywa_B_II_3/Aktywa_B_II_3_A/Aktywa_B_II_3_A_1",
        ),
        "1. REPORTED FIGURES",
        "2. ADJUSTMENTS ",
        confidence_label="Receivables",
    ),
    Mapping(
        "Trade creditors / payables",
        (
            "Bilans/Pasywa/Pasywa_B/Pasywa_B_III/Pasywa_B_III_3/Pasywa_B_III_3_D",
            "Bilans/Pasywa/Pasywa_B/Pasywa_B_III/Pasywa_B_III_3/Pasywa_B_III_3_D/Pasywa_B_III_3_D_1",
        ),
        "1. REPORTED FIGURES",
        "2. ADJUSTMENTS ",
        confidence_label="Payables",
    ),
    Mapping(
        "Cash and cash equivalents",
        (
            "Bilans/Aktywa/Aktywa_B/Aktywa_B_III/Aktywa_B_III_1/Aktywa_B_III_1_C",
            "Bilans/Aktywa/Aktywa_B/Aktywa_B_III",
            "RachPrzeplywow/PrzeplywyPosr/G",
        ),
        "1. REPORTED FIGURES",
        "2. ADJUSTMENTS ",
        confidence_label="Cash & cash equivalents",
    ),
    Mapping(
        "CAPEX",
        ("RachPrzeplywow/PrzeplywyPosr/B/B_II/B_II_1",),
        "1. REPORTED FIGURES",
        "2. ADJUSTMENTS ",
        confidence_label="CAPEX",
    ),
    # Scratchpad debt split
    Mapping(
        "Credits and loans - LT",
        ("Bilans/Pasywa/Pasywa_B/Pasywa_B_II/Pasywa_B_II_3/Pasywa_B_II_3_A",),
        "Interest bearing debt / gross debt",
        "Net revenue",
        confidence_label="Interest-bearing debt",
    ),
    Mapping(
        "Other financial liabilities - LT",
        ("Bilans/Pasywa/Pasywa_B/Pasywa_B_II/Pasywa_B_II_3/Pasywa_B_II_3_C",),
        "Interest bearing debt / gross debt",
        "Net revenue",
        confidence_label="Interest-bearing debt",
    ),
    Mapping(
        "Credits and loans - ST",
        ("Bilans/Pasywa/Pasywa_B/Pasywa_B_III/Pasywa_B_III_3/Pasywa_B_III_3_A",),
        "Interest bearing debt / gross debt",
        "Net revenue",
        confidence_label="Interest-bearing debt",
    ),
    Mapping(
        "Other financial liabilities - ST",
        ("Bilans/Pasywa/Pasywa_B/Pasywa_B_III/Pasywa_B_III_3/Pasywa_B_III_3_C",),
        "Interest bearing debt / gross debt",
        "Net revenue",
        confidence_label="Interest-bearing debt",
    ),
    # Scratchpad revenue split
    Mapping(
        "Sales of products",
        ("RZiS/RZiSKalk/A/A_I", "RZiS/RZiSPor/A/A_I", "RZiS/RZiSKalk/A", "RZiS/RZiSPor/A"),
        "Title",
        "Annualisation",
        row_label_fallbacks=("[Line 1]",),
        write_label="Sales of products",
        confidence_label="Revenue",
    ),
    Mapping(
        "Sales of goods and materials",
        ("RZiS/RZiSKalk/A/A_II", "RZiS/RZiSPor/A/A_IV"),
        "Title",
        "Annualisation",
        row_label_fallbacks=("[Line 2]",),
        write_label="Sales of goods and materials",
        confidence_label="Revenue",
    ),
)


MANUAL_FIELDS = (
    "FTEs (or employees if n/a): not present as a structured XML amount in the training file.",
)


class XmlFinancials:
    def __init__(self, xml_path: Path):
        self.xml_path = xml_path
        self.root = ET.parse(xml_path).getroot()
        self.values: dict[str, dict[str, Decimal]] = {}
        self.period_end = self._first_text("OkresDo")
        self.period_start = self._first_text("OkresOd")
        self.company = self._first_text("NazwaFirmy")
        self._index_amounts()

    @property
    def year(self) -> int:
        if not self.period_end:
            raise ValueError("Could not find OkresDo in XML; pass --target-year explicitly.")
        return datetime.fromisoformat(self.period_end).year

    @property
    def period_start_date(self) -> date | None:
        return date.fromisoformat(self.period_start) if self.period_start else None

    @property
    def period_end_date(self) -> date | None:
        return date.fromisoformat(self.period_end) if self.period_end else None

    @property
    def period_days(self) -> int | None:
        if self.period_start_date is None or self.period_end_date is None:
            return None
        return (self.period_end_date - self.period_start_date).days + 1

    @property
    def requires_annualisation(self) -> bool:
        return (
            self.period_days is not None
            and self.period_days > MIN_ANNUALISATION_DAYS
            and self.period_days not in FULL_YEAR_DAY_COUNTS
        )

    @property
    def comparative_period_tags(self) -> tuple[str, ...]:
        if self.has_nonzero_amount(TRANSFORMED_COMPARATIVE_TAG):
            return COMPARATIVE_PERIOD_TAGS
        return ("KwotaB",)

    def has_nonzero_amount(self, period_tag: str) -> bool:
        return any(
            value is not None and value != 0
            for values in self.values.values()
            for tag, value in values.items()
            if tag == period_tag
        )

    def amount(
        self, paths: Iterable[str], period_tags: Iterable[str]
    ) -> tuple[Decimal | None, str | None, str | None]:
        for path in paths:
            values = self.values.get(path, {})
            for period_tag in period_tags:
                value = values.get(period_tag)
                if value is not None:
                    return value, path, period_tag
        return None, None, None

    def _index_amounts(self) -> None:
        statement_root = self._statement_root()

        def walk(element: ET.Element, parents: list[str]) -> None:
            name = local_name(element.tag)
            path = "/".join(parents + [name]) if parents else name
            direct_values: dict[str, Decimal] = {}
            for child in list(element):
                child_name = local_name(child.tag)
                if child_name in AMOUNT_TAGS and child.text:
                    try:
                        direct_values[child_name] = Decimal(child.text.strip())
                    except InvalidOperation:
                        pass
            if direct_values:
                self.values[path] = direct_values
            for child in list(element):
                child_name = local_name(child.tag)
                if child_name.startswith("Signature") or child_name in AMOUNT_TAGS:
                    continue
                walk(child, parents + [name] if parents else [name])

        for child in list(statement_root):
            child_name = local_name(child.tag)
            if child_name.startswith("Signature"):
                continue
            walk(child, [])

    def _statement_root(self) -> ET.Element:
        if self._has_statement_sections(self.root):
            return self.root

        for element in self.root.iter():
            if local_name(element.tag) in STATEMENT_ROOT_TAGS and self._has_statement_sections(element):
                return element

        for element in self.root.iter():
            if self._has_statement_sections(element):
                return element

        return self.root

    @staticmethod
    def _has_statement_sections(element: ET.Element) -> bool:
        return any(local_name(child.tag) in STATEMENT_SECTION_TAGS for child in list(element))

    def _first_text(self, tag_name: str) -> str | None:
        for element in self.root.iter():
            if local_name(element.tag) == tag_name and element.text:
                return element.text.strip()
        return None


def local_name(tag: str) -> str:
    return tag.rsplit("}", 1)[-1] if "}" in tag else tag


def normalize(value: object) -> str:
    if value is None:
        return ""
    return re.sub(r"\s+", " ", str(value)).strip().lower()


def find_year_column(ws, year: int) -> int:
    for cell in ws[HEADER_ROW]:
        if cell.value == year or str(cell.value).strip() == str(year):
            return cell.column
    raise ValueError(f"Could not find year {year} in row {HEADER_ROW} of the Financials sheet.")


def header_year(value: object) -> int | None:
    if isinstance(value, int):
        return value
    if isinstance(value, float) and value.is_integer():
        return int(value)
    if isinstance(value, str) and re.fullmatch(r"\d{4}", value.strip()):
        return int(value.strip())
    return None


def format_column_runs(columns: Iterable[int]) -> str:
    sorted_columns = sorted(set(columns))
    if not sorted_columns:
        return "none"

    ranges: list[str] = []
    start = previous = sorted_columns[0]
    for column in sorted_columns[1:]:
        if column == previous + 1:
            previous = column
            continue
        ranges.append(column_range_label(start, previous))
        start = previous = column
    ranges.append(column_range_label(start, previous))
    return ", ".join(ranges)


def column_range_label(start: int, end: int) -> str:
    start_letter = get_column_letter(start)
    end_letter = get_column_letter(end)
    return start_letter if start == end else f"{start_letter}:{end_letter}"


def group_unused_year_columns(ws, jobs: list[FillJob], dry_run: bool) -> list[str]:
    used_years = {job.year for job in jobs}
    year_columns = [
        (cell.column, year)
        for cell in ws[HEADER_ROW]
        if cell.column not in PRESERVED_COLUMNS
        for year in [header_year(cell.value)]
        if year is not None
    ]
    if not year_columns:
        return ["SKIP year column grouping: no year columns found"]

    used_columns: list[int] = []
    unused_columns: list[int] = []
    hidden_runs: list[tuple[int, int]] = []
    run_start: int | None = None
    run_end: int | None = None

    for column, year in year_columns:
        column_letter = get_column_letter(column)
        dimension = ws.column_dimensions[column_letter]

        if year in used_years:
            used_columns.append(column)
            if run_start is not None and run_end is not None:
                hidden_runs.append((run_start, run_end))
                run_start = run_end = None
            if not dry_run:
                dimension.hidden = False
                dimension.outlineLevel = 0
                dimension.collapsed = False
            continue

        unused_columns.append(column)
        if run_start is None:
            run_start = column
        run_end = column
        if not dry_run:
            dimension.hidden = True
            dimension.outlineLevel = 1
            dimension.collapsed = False

    if run_start is not None and run_end is not None:
        hidden_runs.append((run_start, run_end))

    if not dry_run:
        visible_year_columns = set(used_columns)
        for _, run_end in hidden_runs:
            for column, _ in year_columns:
                if column > run_end and column in visible_year_columns:
                    ws.column_dimensions[get_column_letter(column)].collapsed = True
                    break

    return [
        f"SHOW year columns: {format_column_runs(used_columns)}",
        f"GROUP hidden unused year columns: {format_column_runs(unused_columns)}",
    ]


def row_label(ws, row: int) -> str:
    for col in (4, 3, 2):
        value = ws.cell(row=row, column=col).value
        if value not in (None, ""):
            return str(value)
    return ""


def find_row(ws, label: str, section_after: str | None, section_before: str | None) -> int:
    label_norm = normalize(label)
    after_row = 1
    before_row = ws.max_row + 1

    if section_after:
        section_norm = normalize(section_after)
        for row in range(1, ws.max_row + 1):
            if normalize(row_label(ws, row)) == section_norm:
                after_row = row
                break

    if section_before:
        section_norm = normalize(section_before)
        for row in range(after_row + 1, ws.max_row + 1):
            if normalize(row_label(ws, row)) == section_norm:
                before_row = row
                break

    matches = [
        row
        for row in range(after_row + 1, before_row)
        if normalize(row_label(ws, row)) == label_norm
    ]
    if not matches:
        scope = f" after {section_after!r}" if section_after else ""
        raise ValueError(f"Could not find row label {label!r}{scope}.")
    if len(matches) > 1:
        raise ValueError(f"Found multiple rows for label {label!r}: {matches}")
    return matches[0]


def find_mapping_row(ws, mapping: Mapping) -> int:
    labels = (mapping.row_label,) + mapping.row_label_fallbacks
    errors: list[str] = []
    for label in labels:
        try:
            return find_row(ws, label, mapping.section_after, mapping.section_before)
        except ValueError as exc:
            errors.append(str(exc))
    raise ValueError("; ".join(errors))


def find_revenue_split_total_row(ws) -> int:
    try:
        annualisation_row = find_row(ws, "Annualisation", None, None)
    except ValueError as exc:
        raise ValueError("Could not find the revenue split total before Annualisation.") from exc

    candidate_rows: list[int] = []
    for row in range(1, annualisation_row):
        if normalize(row_label(ws, row)) == "total":
            candidate_rows.append(row)

    if not candidate_rows:
        raise ValueError("Could not find the scratchpad revenue total row.")
    return candidate_rows[-1]


def find_debt_component_rows(ws) -> list[int]:
    labels = (
        "Credits and loans - LT",
        "Other financial liabilities - LT",
        "Credits and loans - ST",
        "Other financial liabilities - ST",
    )
    return [
        find_row(ws, label, "Interest bearing debt / gross debt", "Net revenue")
        for label in labels
    ]


def find_confidence_row(ws, label: str) -> int:
    return find_row(ws, label, "2. CONFIDENCE LEVEL", "Inputs and calculations")


def scaled_excel_value(raw_value: Decimal, scale: Decimal, blank_if_zero: bool) -> float | None:
    if blank_if_zero and raw_value == 0:
        return None
    return float(raw_value / scale)


def cell_is_empty(cell) -> bool:
    return cell.value in (None, "")


def group_empty_cogs_rows(ws, jobs: list[FillJob], dry_run: bool) -> list[str]:
    cogs_rows = [
        find_row(ws, label, "1. REPORTED FIGURES", "2. ADJUSTMENTS ")
        for label in ("COGS", "Other COGS")
    ]
    gross_margin_row = find_row(
        ws, "Gross margin", "1. REPORTED FIGURES", "2. ADJUSTMENTS "
    )
    used_columns = [find_year_column(ws, job.year) for job in jobs]
    has_cogs_values = any(
        not cell_is_empty(ws.cell(row=row, column=column))
        for row in cogs_rows
        for column in used_columns
    )

    if not dry_run:
        for row in cogs_rows:
            dimension = ws.row_dimensions[row]
            dimension.hidden = not has_cogs_values
            dimension.outlineLevel = 0 if has_cogs_values else 1
            dimension.collapsed = False

        gross_margin_dimension = ws.row_dimensions[gross_margin_row]
        gross_margin_dimension.hidden = False
        gross_margin_dimension.outlineLevel = 0
        gross_margin_dimension.collapsed = not has_cogs_values

    if has_cogs_values:
        return ["SHOW COGS rows: values found in reported figures"]
    return [f"GROUP hidden empty COGS rows: {cogs_rows[0]}:{cogs_rows[-1]}"]


def set_formula_font(cell) -> None:
    font = copy(cell.font)
    font.color = "000000"
    cell.font = font


def update_cagr_formulas(ws, jobs: list[FillJob], dry_run: bool) -> list[str]:
    years = sorted({job.year for job in jobs})
    if len(years) < 2:
        return ["SKIP U CAGR formulas: fewer than two years planned"]

    cagr_years = years[-5:]
    start_year = cagr_years[0]
    end_year = cagr_years[-1]
    start_col = find_year_column(ws, start_year)
    end_col = find_year_column(ws, end_year)
    start_letter = get_column_letter(start_col)
    end_letter = get_column_letter(end_col)
    cagr_letter = get_column_letter(CAGR_COLUMN)

    messages: list[str] = []
    updated = 0
    for row in range(1, ws.max_row + 1):
        cell = ws.cell(row=row, column=CAGR_COLUMN)
        if not (isinstance(cell.value, str) and cell.value.startswith("=")):
            continue

        formula = (
            f'=IFERROR(({end_letter}{row}/{start_letter}{row})^'
            f'(1/({end_letter}${HEADER_ROW}-{start_letter}${HEADER_ROW}))-1,"")'
        )
        if not dry_run:
            cell.value = formula
        updated += 1

    header_cell = ws.cell(row=HEADER_ROW, column=CAGR_COLUMN)
    if isinstance(header_cell.value, str) and normalize(header_cell.value).startswith("cagr"):
        if not dry_run:
            header_cell.value = f"CAGR {start_year}-{end_year}"

    messages.append(
        f"LINK {cagr_letter} CAGR formulas: {start_year}-{end_year} "
        f"({start_letter}:{end_letter}) across {updated} existing formula cells"
    )
    return messages


def source_comment(xml_data: XmlFinancials, period_tag: str, xml_path: str) -> Comment:
    return Comment(f"Gain:\nAR{xml_data.year}", "Codex")


def find_annualisation_row(ws, label: str) -> int:
    return find_row(ws, label, "Annualisation", None)


def set_formula_cell(ws, row: int, col: int, formula: str, dry_run: bool) -> None:
    if not dry_run:
        cell = ws.cell(row=row, column=col)
        cell.value = formula
        cell.comment = None
        set_formula_font(cell)


def configure_annualisation(
    ws,
    period_xml_data: XmlFinancials,
    year: int,
    col: int,
    dry_run: bool,
) -> tuple[dict[str, int], list[str]]:
    col_letter = get_column_letter(col)
    annualisation_start_row = find_row(ws, "Annualisation", None, None)
    annualisation_end_row = find_annualisation_row(ws, "YoY growth (%)")
    rows = {
        label: find_annualisation_row(ws, label)
        for label in (
            "Starting date",
            "Ending date",
            "Annualisation factor",
            "Footnote (copy on CMS)",
            "Net revenue",
            "Other income",
            "Revenue",
            "COGS",
            "Other COGS",
            "Gross margin",
            "Reported EBIT",
            "Reported EBIT Check",
            "Total depreciation",
            "Total amortisation",
            "D&A",
            "EBITDA",
            "CAPEX",
        )
    }
    if period_xml_data.period_start_date is None or period_xml_data.period_end_date is None:
        raise ValueError(f"Could not identify the reporting period dates for {year}.")

    if not dry_run:
        for row in range(annualisation_start_row, annualisation_end_row + 1):
            ws.row_dimensions[row].hidden = False
        start_cell = ws.cell(row=rows["Starting date"], column=col)
        end_cell = ws.cell(row=rows["Ending date"], column=col)
        start_cell.value = period_xml_data.period_start_date
        end_cell.value = period_xml_data.period_end_date
        start_cell.number_format = "dd-mm-yyyy"
        end_cell.number_format = "dd-mm-yyyy"

    set_formula_cell(
        ws,
        rows["Annualisation factor"],
        col,
        (
            f'=IF(AND({col_letter}{rows["Starting date"]}<>0,'
            f'{col_letter}{rows["Ending date"]}<>0),'
            f'({col_letter}{rows["Ending date"]}-{col_letter}{rows["Starting date"]}+1)/365,"")'
        ),
        dry_run,
    )
    set_formula_cell(
        ws,
        rows["Footnote (copy on CMS)"],
        col,
        (
            f'=IF({col_letter}{rows["Annualisation factor"]}<>"","FY"&{col_letter}${HEADER_ROW}'
            f'&": the Company reported for the period from "&TEXT({col_letter}{rows["Starting date"]},"dd-mm")'
            f'&" to "&TEXT({col_letter}{rows["Ending date"]},"dd-mm")'
            f'&". Figures have been annualised to reflect the full-year effect, assuming no seasonality.","")'
        ),
        dry_run,
    )
    scratchpad_formulas = {
        "Revenue": (
            f'=IFERROR({col_letter}{rows["Net revenue"]}+'
            f'{col_letter}{rows["Other income"]},"")'
        ),
        "Gross margin": (
            f'={col_letter}{rows["Revenue"]}-IFERROR({col_letter}{rows["COGS"]}+'
            f'{col_letter}{rows["Other COGS"]},"")'
        ),
        "Reported EBIT Check": (
            f'={col_letter}{rows["EBITDA"]}-{col_letter}{rows["Total depreciation"]}'
            f'-{col_letter}{rows["Total amortisation"]}'
        ),
        "EBITDA": (
            f'=IFERROR({col_letter}{rows["D&A"]}+{col_letter}{rows["Reported EBIT"]},"")'
        ),
    }
    for label, formula in scratchpad_formulas.items():
        set_formula_cell(ws, rows[label], col, formula, dry_run)

    reported_rows = {
        label: find_row(ws, label, "1. REPORTED FIGURES", "2. ADJUSTMENTS ")
        for label in (
            "Net revenue",
            "Other income",
            "Revenue",
            "COGS",
            "Other COGS",
            "Gross margin",
            "Reported EBIT",
            "D&A",
            "EBITDA",
            "CAPEX",
        )
    }
    reported_formulas = {
        "Revenue": (
            f'=IFERROR({col_letter}{reported_rows["Net revenue"]}+'
            f'{col_letter}{reported_rows["Other income"]},"")'
        ),
        "Gross margin": (
            f'={col_letter}{reported_rows["Revenue"]}-IFERROR({col_letter}{reported_rows["COGS"]}+'
            f'{col_letter}{reported_rows["Other COGS"]},"")'
        ),
        "EBITDA": (
            f'=IFERROR({col_letter}{reported_rows["D&A"]}+{col_letter}{reported_rows["Reported EBIT"]},"")'
        ),
    }
    for label, formula in reported_formulas.items():
        set_formula_cell(ws, reported_rows[label], col, formula, dry_run)

    messages = [
        (
            f"ANNUALISE {year}: {period_xml_data.period_start_date.isoformat()} to "
            f"{period_xml_data.period_end_date.isoformat()} ({period_xml_data.period_days} days)"
        ),
        (
            f"SET  {col_letter}{rows['Starting date']}:{col_letter}{rows['Ending date']} "
            "Annualisation dates and factor"
        ),
        f"SHOW annualisation rows: {annualisation_start_row}:{annualisation_end_row}",
    ]
    return rows, messages


def fill_period(
    ws,
    xml_data: XmlFinancials,
    year: int,
    period_tags: tuple[str, ...],
    overwrite: bool,
    add_comments: bool,
    dry_run: bool,
    period_xml_data: XmlFinancials | None = None,
) -> list[str]:
    col = find_year_column(ws, year)
    col_letter = get_column_letter(col)
    messages: list[str] = []
    confidence_labels: set[str] = set()
    net_revenue_row = find_row(ws, "Net revenue", "1. REPORTED FIGURES", "2. ADJUSTMENTS ")
    revenue_total_row = find_revenue_split_total_row(ws)
    interest_bearing_debt_row = find_row(
        ws, "Interest bearing debt / gross debt", "1. REPORTED FIGURES", "2. ADJUSTMENTS "
    )
    debt_component_rows = find_debt_component_rows(ws)
    annualisation_rows: dict[str, int] | None = None
    if period_xml_data and period_xml_data.requires_annualisation:
        annualisation_rows, annualisation_messages = configure_annualisation(
            ws, period_xml_data, year, col, dry_run
        )
        messages.extend(annualisation_messages)

    for mapping in MAPPINGS:
        row = find_mapping_row(ws, mapping)
        input_row = (
            annualisation_rows[mapping.row_label]
            if annualisation_rows and mapping.row_label in {"Other income", "COGS", "Reported EBIT", "D&A", "CAPEX"}
            else row
        )
        cell = ws.cell(row=input_row, column=col)
        raw_value, xml_path, period_tag = xml_data.amount(mapping.xml_paths, period_tags)

        if raw_value is None or xml_path is None or period_tag is None:
            messages.append(f"MISS {col_letter}{input_row} {mapping.row_label}: no XML value found")
            continue

        if mapping.confidence_label:
            confidence_labels.add(mapping.confidence_label)

        value = scaled_excel_value(raw_value, mapping.scale, mapping.blank_if_zero)
        if not overwrite and not cell_is_empty(cell):
            messages.append(f"SKIP {col_letter}{row} {mapping.row_label}: cell already populated")
            continue

        if not dry_run:
            if mapping.write_label:
                ws.cell(row=row, column=4).value = mapping.write_label
            cell.value = value
            if add_comments and value is not None:
                cell.comment = source_comment(xml_data, period_tag, xml_path)
            if annualisation_rows and input_row != row:
                reported_cell = ws.cell(row=row, column=col)
                reported_cell.value = (
                    f"={col_letter}{input_row}/${col_letter}${annualisation_rows['Annualisation factor']}"
                )
                reported_cell.comment = None
                set_formula_font(reported_cell)
        rendered = "" if value is None else f"{value:.8f}".rstrip("0").rstrip(".")
        messages.append(f"SET  {col_letter}{input_row} {mapping.row_label}: {rendered} from {xml_path}/{period_tag}")
        if annualisation_rows and input_row != row:
            messages.append(
                f"LINK {col_letter}{row} {mapping.row_label}: "
                f"={col_letter}{input_row}/${col_letter}${annualisation_rows['Annualisation factor']}"
            )

    net_revenue_cell = ws.cell(row=net_revenue_row, column=col)
    if annualisation_rows:
        annual_revenue_cell = ws.cell(row=annualisation_rows["Net revenue"], column=col)
        if overwrite or cell_is_empty(annual_revenue_cell):
            if not dry_run:
                annual_revenue_cell.value = f"={col_letter}{revenue_total_row}"
                annual_revenue_cell.comment = None
                set_formula_font(annual_revenue_cell)
                net_revenue_cell.value = (
                    f"={col_letter}{revenue_total_row}/"
                    f"${col_letter}${annualisation_rows['Annualisation factor']}"
                )
                net_revenue_cell.comment = None
                set_formula_font(net_revenue_cell)
            messages.append(
                f"LINK {col_letter}{annualisation_rows['Net revenue']} Net revenue: "
                f"={col_letter}{revenue_total_row}"
            )
            messages.append(
                f"LINK {col_letter}{net_revenue_row} Net revenue: "
                f"={col_letter}{revenue_total_row}/"
                f"${col_letter}${annualisation_rows['Annualisation factor']}"
            )
        else:
            messages.append(
                f"SKIP {col_letter}{annualisation_rows['Net revenue']} Net revenue: cell already populated"
            )
    elif overwrite or cell_is_empty(net_revenue_cell):
        if not dry_run:
            net_revenue_cell.value = f"={col_letter}{revenue_total_row}"
            net_revenue_cell.comment = None
            set_formula_font(net_revenue_cell)
        messages.append(f"LINK {col_letter}{net_revenue_row} Net revenue: ={col_letter}{revenue_total_row}")
    else:
        messages.append(f"SKIP {col_letter}{net_revenue_row} Net revenue: cell already populated")

    interest_bearing_debt_cell = ws.cell(row=interest_bearing_debt_row, column=col)
    debt_formula = f"=SUM({','.join(f'{col_letter}{row}' for row in debt_component_rows)})"
    if overwrite or cell_is_empty(interest_bearing_debt_cell):
        if not dry_run:
            interest_bearing_debt_cell.value = debt_formula
            interest_bearing_debt_cell.comment = None
            set_formula_font(interest_bearing_debt_cell)
        messages.append(f"LINK {col_letter}{interest_bearing_debt_row} Interest-bearing debt: {debt_formula}")
    else:
        messages.append(f"SKIP {col_letter}{interest_bearing_debt_row} Interest-bearing debt: cell already populated")

    for confidence_label in sorted(confidence_labels):
        try:
            confidence_row = find_confidence_row(ws, confidence_label)
        except ValueError:
            messages.append(f"MISS {col_letter} confidence {confidence_label}: row not found")
            continue

        confidence_cell = ws.cell(row=confidence_row, column=col)
        if overwrite or cell_is_empty(confidence_cell):
            if not dry_run:
                confidence_cell.value = "Actual"
            messages.append(f"CONF {col_letter}{confidence_row} {confidence_label}: Actual")
        else:
            messages.append(
                f"SKIP {col_letter}{confidence_row} {confidence_label} confidence: cell already populated"
            )

    return messages


def save_workbook(wb, output_path: Path) -> None:
    try:
        wb.calculation.fullCalcOnLoad = True
        wb.calculation.forceFullCalc = True
        wb.calculation.calcMode = "auto"
    except AttributeError:
        pass
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)


def build_fill_jobs(
    xml_files: list[XmlFinancials],
    target_year: int | None,
    fill_comparative: bool,
    years: int,
) -> list[FillJob]:
    if years < 1:
        raise ValueError("--years must be at least 1.")

    if len(xml_files) == 1:
        xml_data = xml_files[0]
        current_year = target_year or xml_data.year
        jobs = [FillJob(xml_data, current_year, (CURRENT_PERIOD_TAG,), xml_data)]
        if fill_comparative and years > 1:
            jobs.append(FillJob(xml_data, current_year - 1, xml_data.comparative_period_tags))
        return jobs[:years]

    if target_year is not None:
        raise ValueError("--target-year is only supported when filling from one XML file.")

    sorted_xmls = sorted(xml_files, key=lambda item: item.year, reverse=True)
    period_xml_by_year = {xml_data.year: xml_data for xml_data in sorted_xmls}
    jobs: list[FillJob] = []
    seen_years: set[int] = set()

    for index, xml_data in enumerate(sorted_xmls):
        planned = (
            (
                (xml_data.year, (CURRENT_PERIOD_TAG,)),
                (xml_data.year - 1, xml_data.comparative_period_tags),
            )
            if index == 0
            else ((xml_data.year - 1, xml_data.comparative_period_tags),)
        )
        for year, period_tags in planned:
            if year in seen_years:
                continue
            jobs.append(FillJob(xml_data, year, period_tags, period_xml_by_year.get(year)))
            seen_years.add(year)
            if len(jobs) >= years:
                return jobs

    return jobs


def parse_args(argv: list[str]) -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Fill a Gain.pro-style Financials sheet from a Polish ESF XML file."
    )
    parser.add_argument(
        "--xml",
        required=True,
        nargs="+",
        type=Path,
        help=(
            "Input ESF XML file(s). With multiple XMLs, pass the recent annual filings; "
            "the script sorts them by fiscal year, uses the newest XML's current year "
            "from KwotaA and comparative years from KwotaB1 when present, falling back "
            "to KwotaB when KwotaB1 is missing or all zeros."
        ),
    )
    parser.add_argument("--template", required=True, type=Path, help="Input Excel template/workbook.")
    parser.add_argument("--output", required=True, type=Path, help="Output .xlsx path.")
    parser.add_argument(
        "--target-year",
        type=int,
        help="Year column to fill. Defaults to the XML OkresDo year.",
    )
    parser.add_argument(
        "--fill-comparative",
        action="store_true",
        help=(
            "For a single XML, also fill the previous-year column from KwotaB1 when "
            "present, falling back to KwotaB when KwotaB1 is missing or all zeros."
        ),
    )
    parser.add_argument(
        "--years",
        type=int,
        default=6,
        help="Maximum number of fiscal years to fill when multiple XMLs are provided. Default: 6.",
    )
    parser.add_argument(
        "--no-overwrite",
        action="store_true",
        help="Do not overwrite existing values in mapped cells.",
    )
    parser.add_argument("--no-comments", action="store_true", help="Do not add source comments.")
    parser.add_argument("--dry-run", action="store_true", help="Print planned changes without saving.")
    return parser.parse_args(argv)


def main(argv: list[str]) -> int:
    args = parse_args(argv)
    xml_files = [XmlFinancials(xml_path) for xml_path in args.xml]

    wb = load_workbook(args.template)
    if FINANCIALS_SHEET not in wb.sheetnames:
        raise SystemExit(f"Workbook does not contain a {FINANCIALS_SHEET!r} sheet.")
    ws = wb[FINANCIALS_SHEET]

    try:
        jobs = build_fill_jobs(
            xml_files,
            target_year=args.target_year,
            fill_comparative=args.fill_comparative,
            years=args.years,
        )
    except ValueError as exc:
        raise SystemExit(str(exc)) from exc

    print("XML files:")
    for xml_data in xml_files:
        print(f"- {xml_data.xml_path} ({xml_data.year})")
    print(f"Workbook: {args.template}")
    if xml_files:
        print(f"Company: {xml_files[0].company or 'n/a'}")
    print("Fill plan:")
    for job in jobs:
        print(f"- {job.year}: {job.xml_data.xml_path.name} / {' then '.join(job.period_tags)}")
    print()

    overwrite = not args.no_overwrite
    all_messages: list[str] = []
    for job in jobs:
        all_messages.append(
            f"=== Filling {job.year} from {job.xml_data.xml_path.name} / "
            f"{' then '.join(job.period_tags)} ==="
        )
        all_messages.extend(
            fill_period(
                ws,
                job.xml_data,
                job.year,
                job.period_tags,
                overwrite=overwrite,
                add_comments=not args.no_comments,
                dry_run=args.dry_run,
                period_xml_data=job.period_xml_data,
            )
        )

    all_messages.extend(update_cagr_formulas(ws, jobs, dry_run=args.dry_run))
    all_messages.extend(group_unused_year_columns(ws, jobs, dry_run=args.dry_run))
    all_messages.extend(group_empty_cogs_rows(ws, jobs, dry_run=args.dry_run))

    for message in all_messages:
        print(message)

    print()
    print("Manual / unresolved fields:")
    for field in MANUAL_FIELDS:
        print(f"- {field}")

    if args.dry_run:
        print("\nDry run only; no workbook saved.")
        return 0

    save_workbook(wb, args.output)
    print(f"\nSaved: {args.output}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main(sys.argv[1:]))
