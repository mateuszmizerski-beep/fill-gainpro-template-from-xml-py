#!/usr/bin/env python3
"""Update an existing Gain.pro financials workbook with a new ESF XML filing.

The update flow starts from a fresh template, migrates historical/manual cells
from an existing filled workbook, then overwrites the newest year and prior-year
comparative column from the uploaded XML.
"""

from __future__ import annotations

import argparse
import re
import unicodedata
from copy import copy
from pathlib import Path
from types import SimpleNamespace
from typing import Iterable

from openpyxl import load_workbook
from openpyxl.cell.cell import Cell, MergedCell
from openpyxl.comments import Comment
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from fill_gainpro_template_from_xml import (
    CURRENT_PERIOD_TAG,
    FINANCIALS_SHEET,
    HEADER_ROW,
    XmlFinancials,
    build_fill_jobs,
    cell_is_empty,
    fill_period,
    find_row,
    group_empty_cogs_rows,
    group_unused_year_columns,
    header_year,
    normalize,
    save_workbook,
    update_cagr_formulas,
)


CONTEXTS: tuple[tuple[str, str | None], ...] = (
    ("2. CONFIDENCE LEVEL", "Inputs and calculations"),
    ("1. REPORTED FIGURES", "2. ADJUSTMENTS "),
    ("2. ADJUSTMENTS ", "3. ADJUSTED FIGURES"),
    ("3. ADJUSTED FIGURES", "Scratchpad"),
    ("Scratchpad", "Interest bearing debt / gross debt"),
    ("Interest bearing debt / gross debt", "Annualisation"),
    ("Title", "Annualisation"),
    ("Annualisation", None),
)

IGNORED_SUBSTANTIVE_STRINGS = {
    "actual",
    "estimate",
    "please check comment boxes for sources",
}

IGNORED_COMPANY_TOKENS = {
    "financials",
    "gain",
    "gainpro",
    "pl",
    "poland",
}


def copy_cell_style_and_metadata(source: Cell, destination: Cell) -> None:
    if source.has_style:
        destination.font = copy(source.font)
        destination.fill = copy(source.fill)
        destination.border = copy(source.border)
        destination.alignment = copy(source.alignment)
        destination.protection = copy(source.protection)
        destination.number_format = source.number_format
    if source.comment:
        destination.comment = copy(source.comment)
    else:
        destination.comment = None
    if source.hyperlink:
        destination._hyperlink = copy(source.hyperlink)


def copy_cell(source: Cell, destination: Cell, *, copy_formula: bool) -> None:
    if isinstance(destination, MergedCell):
        return
    if copy_formula or not is_formula(source.value):
        destination.value = source.value
    copy_cell_style_and_metadata(source, destination)


def copy_row_and_column_dimensions(source: Worksheet, destination: Worksheet) -> None:
    for row_idx, dimension in source.row_dimensions.items():
        target = destination.row_dimensions[row_idx]
        target.height = dimension.height
        target.hidden = dimension.hidden
        target.outlineLevel = dimension.outlineLevel
        target.collapsed = dimension.collapsed

    for col_key, dimension in source.column_dimensions.items():
        target = destination.column_dimensions[col_key]
        target.width = dimension.width
        target.hidden = dimension.hidden
        target.outlineLevel = dimension.outlineLevel
        target.collapsed = dimension.collapsed


def copy_sheet_contents(source: Worksheet, destination: Worksheet) -> None:
    for merged_range in list(destination.merged_cells.ranges):
        destination.unmerge_cells(str(merged_range))
    for merged_range in source.merged_cells.ranges:
        destination.merge_cells(str(merged_range))

    for row in source.iter_rows():
        for source_cell in row:
            destination_cell = destination.cell(source_cell.row, source_cell.column)
            if isinstance(destination_cell, MergedCell):
                continue
            copy_cell(source_cell, destination_cell, copy_formula=True)

    copy_row_and_column_dimensions(source, destination)
    destination.freeze_panes = source.freeze_panes
    destination.sheet_view.showGridLines = source.sheet_view.showGridLines


def copy_segments_sheet(source_wb, destination_wb) -> list[str]:
    if "Segments" not in source_wb.sheetnames:
        return ["SKIP Segments sheet: not present in existing workbook"]

    source_ws = source_wb["Segments"]
    if "Segments" in destination_wb.sheetnames:
        index = destination_wb.sheetnames.index("Segments")
        destination_wb.remove(destination_wb["Segments"])
        destination_ws = destination_wb.create_sheet("Segments", index)
    else:
        destination_ws = destination_wb.create_sheet("Segments")

    copy_sheet_contents(source_ws, destination_ws)
    return ["COPY Segments sheet"]


def is_formula(value: object) -> bool:
    return isinstance(value, str) and value.startswith("=")


def plain_text(value: str) -> str:
    without_accents = unicodedata.normalize("NFKD", value)
    return "".join(char for char in without_accents if not unicodedata.combining(char))


def company_tokens(value: str) -> set[str]:
    tokens = {
        token
        for token in re.split(r"[^a-z0-9]+", plain_text(value).lower())
        if len(token) >= 3
    }
    return tokens - IGNORED_COMPANY_TOKENS


def validate_company_match(existing_path: Path, xml_data: XmlFinancials) -> None:
    filename_tokens = company_tokens(existing_path.stem)
    xml_tokens = company_tokens(xml_data.company or "")
    if not filename_tokens or not xml_tokens:
        return
    if filename_tokens.isdisjoint(xml_tokens):
        raise ValueError(
            "Existing workbook filename and XML company name do not appear to match. "
            f"Workbook: {existing_path.name}. XML company: {xml_data.company or 'n/a'}. "
            "Pass --allow-company-mismatch to override."
        )


def is_substantive_value(value: object) -> bool:
    if value in (None, "") or is_formula(value):
        return False
    if isinstance(value, str) and normalize(value) in IGNORED_SUBSTANTIVE_STRINGS:
        return False
    return True


def financial_year_columns(ws: Worksheet) -> dict[int, int]:
    columns: dict[int, int] = {}
    for cell in ws[HEADER_ROW]:
        year = header_year(cell.value)
        if year is not None:
            columns[year] = cell.column
    return columns


def active_years(ws: Worksheet) -> list[int]:
    years: list[int] = []
    for year, column in financial_year_columns(ws).items():
        substantive_cells = 0
        for row in range(1, ws.max_row + 1):
            if is_substantive_value(ws.cell(row=row, column=column).value):
                substantive_cells += 1
            if substantive_cells >= 3:
                years.append(year)
                break
    return sorted(years)


def first_label_after(ws: Worksheet, label: str, after_row: int = 0) -> int | None:
    label_norm = normalize(label)
    for row in range(after_row + 1, ws.max_row + 1):
        if normalize(source_row_label(ws, row)) == label_norm:
            return row
    return None


def context_ranges(ws: Worksheet) -> list[tuple[int, int, str, str | None]]:
    confidence = first_label_after(ws, "2. CONFIDENCE LEVEL")
    inputs = first_label_after(ws, "Inputs and calculations", confidence or 0)
    reported = first_label_after(ws, "1. REPORTED FIGURES")
    adjustments = first_label_after(ws, "2. ADJUSTMENTS ", reported or 0)
    adjusted = first_label_after(ws, "3. ADJUSTED FIGURES", adjustments or 0)
    scratchpad = first_label_after(ws, "Scratchpad", adjusted or 0)
    debt = first_label_after(ws, "Interest bearing debt / gross debt", scratchpad or 0)
    title = first_label_after(ws, "Title", debt or scratchpad or 0)
    annualisation = first_label_after(ws, "Annualisation", scratchpad or 0)

    ranges: list[tuple[int, int, str, str | None]] = []
    if confidence and inputs:
        ranges.append((confidence, inputs, "2. CONFIDENCE LEVEL", "Inputs and calculations"))
    if reported and adjustments:
        ranges.append((reported, adjustments, "1. REPORTED FIGURES", "2. ADJUSTMENTS "))
    if adjustments and adjusted:
        ranges.append((adjustments, adjusted, "2. ADJUSTMENTS ", "3. ADJUSTED FIGURES"))
    if adjusted and scratchpad:
        ranges.append((adjusted, scratchpad, "3. ADJUSTED FIGURES", "Scratchpad"))
    if scratchpad:
        scratchpad_end = debt or annualisation or ws.max_row + 1
        ranges.append((scratchpad, scratchpad_end, "Scratchpad", "Interest bearing debt / gross debt"))
    if debt:
        debt_end = title or annualisation or ws.max_row + 1
        ranges.append((debt, debt_end, "Interest bearing debt / gross debt", "Annualisation"))
    if title and annualisation:
        ranges.append((title, annualisation, "Title", "Annualisation"))
    if annualisation:
        ranges.append((annualisation, ws.max_row + 1, "Annualisation", None))
    return ranges


def row_context(
    ranges: Iterable[tuple[int, int, str, str | None]], row: int
) -> tuple[str, str | None] | None:
    for start, end, section_after, section_before in ranges:
        if start < row < end:
            return section_after, section_before
    return None


def destination_row_for_source(
    source_ws: Worksheet,
    destination_ws: Worksheet,
    source_contexts: list[tuple[int, int, str, str | None]],
    destination_contexts: list[tuple[int, int, str, str | None]],
    source_row: int,
) -> int | None:
    label = source_row_label(source_ws, source_row)
    if not label:
        return None

    context = row_context(source_contexts, source_row)
    if not context:
        return None

    section_after, section_before = context
    for start, end, destination_after, destination_before in destination_contexts:
        if destination_after == section_after and destination_before == section_before:
            return find_label_between(destination_ws, label, start, end)
    return None


def find_label_between(ws: Worksheet, label: str, start: int, end: int) -> int | None:
    label_norm = normalize(label)
    matches = [
        row
        for row in range(start + 1, end)
        if normalize(source_row_label(ws, row)) == label_norm
    ]
    if len(matches) == 1:
        return matches[0]
    return None


def source_row_label(ws: Worksheet, row: int) -> str:
    for col in (4, 3, 2):
        value = ws.cell(row=row, column=col).value
        if value not in (None, ""):
            return str(value)
    return ""


def should_copy_financials_cell(source: Cell) -> bool:
    return source.value not in (None, "") or source.comment is not None or source.has_style


def unique_label_row(ws: Worksheet, *labels: str) -> int | None:
    label_norms = {normalize(label) for label in labels}
    matches = [
        row
        for row in range(1, ws.max_row + 1)
        if normalize(source_row_label(ws, row)) in label_norms
    ]
    return matches[0] if len(matches) == 1 else None


def revenue_migration_rows(
    source_ws: Worksheet,
    destination_ws: Worksheet,
) -> tuple[dict[int, int], int | None, int, int] | None:
    source_products = unique_label_row(source_ws, "Sales of products")
    source_goods = unique_label_row(source_ws, "Sales of goods and materials")
    destination_products = unique_label_row(
        destination_ws, "Sales of products", "[Line 1]"
    )
    destination_goods = unique_label_row(
        destination_ws, "Sales of goods and materials", "[Line 2]"
    )
    if None in (
        source_products,
        source_goods,
        destination_products,
        destination_goods,
    ):
        return None

    source_total = first_label_after(source_ws, "Total", source_goods)
    destination_total = first_label_after(destination_ws, "Total", destination_goods)
    if destination_total is None:
        return None

    source_reported_revenue = find_row(
        source_ws, "Net revenue", "1. REPORTED FIGURES", "2. ADJUSTMENTS "
    )
    destination_reported_revenue = find_row(
        destination_ws, "Net revenue", "1. REPORTED FIGURES", "2. ADJUSTMENTS "
    )
    return (
        {
            source_products: destination_products,
            source_goods: destination_goods,
        },
        source_total,
        source_reported_revenue,
        destination_reported_revenue,
    )


def migrate_financials_columns(
    source_ws: Worksheet,
    destination_ws: Worksheet,
    years: Iterable[int],
) -> list[str]:
    source_year_columns = financial_year_columns(source_ws)
    destination_year_columns = financial_year_columns(destination_ws)
    source_contexts = context_ranges(source_ws)
    destination_contexts = context_ranges(destination_ws)
    destination_row_cache: dict[int, int | None] = {}
    messages: list[str] = []
    revenue_rows = revenue_migration_rows(source_ws, destination_ws)
    revenue_component_rows: dict[int, int] = {}
    source_revenue_total: int | None = None
    source_reported_revenue: int | None = None
    destination_reported_revenue: int | None = None
    if revenue_rows:
        (
            revenue_component_rows,
            source_revenue_total,
            source_reported_revenue,
            destination_reported_revenue,
        ) = revenue_rows
        destination_rows = list(revenue_component_rows.values())
        destination_ws.cell(destination_rows[0], 4).value = "Sales of products"
        destination_ws.cell(destination_rows[1], 4).value = "Sales of goods and materials"

    for year in years:
        if year not in source_year_columns or year not in destination_year_columns:
            messages.append(f"SKIP migrated year {year}: year column missing")
            continue

        source_col = source_year_columns[year]
        destination_col = destination_year_columns[year]
        copied = 0

        for row in range(1, source_ws.max_row + 1):
            if row in revenue_component_rows or row == source_revenue_total:
                continue
            source_cell = source_ws.cell(row=row, column=source_col)
            if not should_copy_financials_cell(source_cell):
                continue

            if row not in destination_row_cache:
                destination_row_cache[row] = destination_row_for_source(
                    source_ws, destination_ws, source_contexts, destination_contexts, row
                )
            destination_row = destination_row_cache[row]
            if destination_row is None:
                continue

            destination_cell = destination_ws.cell(row=destination_row, column=destination_col)
            copy_cell(source_cell, destination_cell, copy_formula=False)
            copied += 1

        if revenue_rows:
            destination_component_rows = list(revenue_component_rows.values())
            for source_row, destination_row in revenue_component_rows.items():
                source_cell = source_ws.cell(row=source_row, column=source_col)
                if should_copy_financials_cell(source_cell):
                    destination_cell = destination_ws.cell(
                        row=destination_row, column=destination_col
                    )
                    copy_cell(source_cell, destination_cell, copy_formula=False)
                    copied += 1

            destination_total = first_label_after(
                destination_ws, "Total", max(destination_component_rows)
            )
            col_letter = get_column_letter(destination_col)
            if destination_total is not None:
                destination_ws.cell(destination_total, destination_col).value = (
                    f"=SUM({col_letter}{destination_component_rows[0]}:"
                    f"{col_letter}{destination_component_rows[1]})"
                )
                reported_cell = destination_ws.cell(
                    destination_reported_revenue, destination_col
                )
                source_reported_cell = source_ws.cell(
                    source_reported_revenue, source_col
                )
                copy_cell_style_and_metadata(source_reported_cell, reported_cell)
                reported_cell.value = f"={col_letter}{destination_total}"
                reported_cell.comment = None

        messages.append(
            f"COPY Financials {year}: {get_column_letter(source_col)} -> "
            f"{get_column_letter(destination_col)} ({copied} cells)"
        )
    return messages


def display_year_jobs(years: Iterable[int]) -> list[SimpleNamespace]:
    return [SimpleNamespace(year=year) for year in sorted(set(years))]


def update_workbook(
    existing_path: Path,
    template_path: Path,
    xml_paths: Iterable[Path],
    output_path: Path,
    *,
    add_comments: bool = True,
    allow_company_mismatch: bool = False,
) -> list[str]:
    xml_files = [XmlFinancials(xml_path) for xml_path in xml_paths]
    if not xml_files:
        raise ValueError("At least one XML file is required.")
    if not allow_company_mismatch:
        for xml_data in xml_files:
            validate_company_match(existing_path, xml_data)

    existing_wb = load_workbook(existing_path)
    output_wb = load_workbook(template_path)

    if FINANCIALS_SHEET not in existing_wb.sheetnames:
        raise ValueError(f"Existing workbook does not contain a {FINANCIALS_SHEET!r} sheet.")
    if FINANCIALS_SHEET not in output_wb.sheetnames:
        raise ValueError(f"Template workbook does not contain a {FINANCIALS_SHEET!r} sheet.")

    messages: list[str] = []
    messages.extend(copy_segments_sheet(existing_wb, output_wb))

    existing_ws = existing_wb[FINANCIALS_SHEET]
    output_ws = output_wb[FINANCIALS_SHEET]
    existing_years = active_years(existing_ws)
    messages.append(f"Existing populated years: {', '.join(map(str, existing_years)) or 'none'}")
    messages.extend(migrate_financials_columns(existing_ws, output_ws, existing_years))

    newest_xml_year = max(xml_data.year for xml_data in xml_files)
    latest_existing_year = max(existing_years) if existing_years else newest_xml_year - 1
    fill_year_count = max(2, newest_xml_year - latest_existing_year)
    fill_jobs = build_fill_jobs(
        xml_files,
        target_year=None,
        fill_comparative=True,
        years=fill_year_count,
    )
    display_years = sorted(set(existing_years) | {job.year for job in fill_jobs})
    messages.append(f"Displayed output years: {', '.join(map(str, display_years))}")

    for job in fill_jobs:
        messages.append(
            f"=== Filling {job.year} from {job.xml_data.xml_path.name} / "
            f"{' then '.join(job.period_tags)} ==="
        )
        messages.extend(
            fill_period(
                output_ws,
                job.xml_data,
                job.year,
                job.period_tags,
                overwrite=True,
                add_comments=add_comments,
                dry_run=False,
                period_xml_data=job.period_xml_data,
            )
        )

    visible_jobs = display_year_jobs(display_years)
    messages.extend(update_cagr_formulas(output_ws, visible_jobs, dry_run=False))
    messages.extend(group_unused_year_columns(output_ws, visible_jobs, dry_run=False))
    messages.extend(group_empty_cogs_rows(output_ws, visible_jobs, dry_run=False))

    save_workbook(output_wb, output_path)
    messages.append(f"Saved: {output_path}")
    return messages


def parse_args(argv: list[str] | None = None) -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Update a filled Gain.pro financials workbook from a new ESF XML."
    )
    parser.add_argument("--existing", required=True, type=Path, help="Existing filled workbook.")
    parser.add_argument(
        "--xml",
        required=True,
        nargs="+",
        type=Path,
        help=(
            "Newest ESF XML filing, or multiple recent XML filings. The newest XML "
            "refreshes both its current year and prior-year comparative values."
        ),
    )
    parser.add_argument("--template", required=True, type=Path, help="Fresh template workbook.")
    parser.add_argument("--output", required=True, type=Path, help="Output workbook path.")
    parser.add_argument("--no-comments", action="store_true", help="Do not add XML source comments.")
    parser.add_argument(
        "--allow-company-mismatch",
        action="store_true",
        help="Allow the XML company name to differ from the existing workbook filename.",
    )
    return parser.parse_args(argv)


def main(argv: list[str] | None = None) -> int:
    args = parse_args(argv)
    try:
        messages = update_workbook(
            args.existing,
            args.template,
            args.xml,
            args.output,
            add_comments=not args.no_comments,
            allow_company_mismatch=args.allow_company_mismatch,
        )
    except Exception as exc:
        raise SystemExit(str(exc)) from exc

    for message in messages:
        print(message)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
